#!/usr/bin/env python3
"""
Extract AidData CLG-LMIC v1.0 Excel workbook into clean relational CSV tables.

Design:
- raw Excel remains immutable
- output is pandas-friendly CSV
- column names are normalized to snake_case
- original column mapping is preserved
- known relational tables are exported separately
- relationship/key audits are produced

Expected workbook sheets:
- CLG-LMIC 1.0_Records
- Borrower Ownership
- CountryList
- Definitions_Records
- Definitions_Borrower Ownership
- Contents / Overview / Guide / License are documentation sheets and skipped by default
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook


KNOWN_TABLES = {
    "CLG-LMIC 1.0_Records": {
        "table_name": "aiddata_records",
        "header": 0,
        "role": "main_project_activity_table",
        "primary_key": "aiddata_record_id",
    },
    "Borrower Ownership": {
        "table_name": "aiddata_borrower_ownership",
        "header": 0,
        "role": "child_borrower_ownership_table",
        "foreign_key": "aiddata_record_id",
    },
    "CountryList": {
        "table_name": "aiddata_country_list",
        "header": 4,
        "role": "country_reference_table",
    },
    "Definitions_Records": {
        "table_name": "aiddata_definitions_records",
        "header": 3,
        "role": "data_dictionary_records",
    },
    "Definitions_Borrower Ownership": {
        "table_name": "aiddata_definitions_borrower_ownership",
        "header": 3,
        "role": "data_dictionary_borrower_ownership",
    },
}

DOC_SHEETS = {"Contents", "Overview", "Guide", "License"}


def clean_column_name(col: object) -> str:
    """Convert arbitrary Excel column name to stable snake_case."""
    c = "" if col is None else str(col).strip()
    c = c.replace("\n", " ")
    c = re.sub(r"\s+", "_", c)
    c = re.sub(r"[^0-9a-zA-Z_]+", "_", c)
    c = re.sub(r"_+", "_", c).strip("_").lower()

    if not c:
        c = "unnamed"

    if re.match(r"^\d", c):
        c = "col_" + c

    return c


def dedupe_columns(cols: List[object]) -> Tuple[List[str], pd.DataFrame]:
    """Clean columns and make duplicates explicit with suffixes."""
    seen: Dict[str, int] = {}
    cleaned: List[str] = []
    rows = []

    for idx, original in enumerate(cols):
        base = clean_column_name(original)

        if base not in seen:
            seen[base] = 0
            final = base
        else:
            seen[base] += 1
            final = f"{base}_{seen[base]}"

        cleaned.append(final)
        rows.append(
            {
                "column_position": idx,
                "original_column": "" if original is None else str(original),
                "clean_column": final,
                "base_clean_column": base,
                "was_duplicate": final != base,
            }
        )

    return cleaned, pd.DataFrame(rows)


def drop_empty_rows_and_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Remove fully empty rows/columns, preserving all source variables otherwise."""
    df = df.copy()
    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")

    # Normalize purely blank strings to NA, then re-drop empty rows/cols.
    df = df.replace(r"^\s*$", pd.NA, regex=True)
    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")

    return df


def read_known_sheet(xlsx_path: Path, sheet_name: str, header: int) -> pd.DataFrame:
    df = pd.read_excel(
        xlsx_path,
        sheet_name=sheet_name,
        header=header,
        dtype=str,
    )
    df = drop_empty_rows_and_cols(df)
    return df


def workbook_sheet_inventory(xlsx_path: Path) -> pd.DataFrame:
    """Lightweight sheet inventory without loading full sheets into pandas."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    rows = []
    for ws in wb.worksheets:
        preview_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            preview_rows.append(" | ".join("" if v is None else str(v) for v in row[:12]))
            if i >= 5:
                break

        rows.append(
            {
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimension": ws.calculate_dimension(force=True),
                "known_role": (
                    KNOWN_TABLES.get(ws.title, {}).get("role")
                    if ws.title in KNOWN_TABLES
                    else ("documentation_sheet" if ws.title in DOC_SHEETS else "unknown")
                ),
                "preview_first_rows": "\n".join(preview_rows),
            }
        )

    return pd.DataFrame(rows)


def export_table(
    df: pd.DataFrame,
    table_name: str,
    sheet_name: str,
    out_tables: Path,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Clean names, add source metadata, write CSV, return cleaned df and mapping."""
    original_cols = list(df.columns)
    clean_cols, mapping = dedupe_columns(original_cols)

    out = df.copy()
    out.columns = clean_cols

    # Keep minimal provenance columns at the end.
    out["source_sheet"] = sheet_name
    out["source_table"] = table_name

    table_path = out_tables / f"{table_name}.csv"
    out.to_csv(table_path, index=False)

    mapping.insert(0, "sheet_name", sheet_name)
    mapping.insert(1, "table_name", table_name)

    return out, mapping


def build_table_inventory(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []

    for table_name, df in tables.items():
        rows.append(
            {
                "table_name": table_name,
                "n_rows": len(df),
                "n_cols": len(df.columns),
                "columns": " | ".join(df.columns),
            }
        )

    return pd.DataFrame(rows)


def build_key_uniqueness_audit(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    checks = []

    def add_check(table_name: str, key: str):
        df = tables.get(table_name)
        if df is None:
            checks.append(
                {
                    "table_name": table_name,
                    "key_column": key,
                    "exists": False,
                    "n_rows": None,
                    "n_nonmissing_key": None,
                    "n_unique_key": None,
                    "n_duplicate_rows_by_key": None,
                    "status": "missing_table",
                }
            )
            return

        exists = key in df.columns
        if not exists:
            checks.append(
                {
                    "table_name": table_name,
                    "key_column": key,
                    "exists": False,
                    "n_rows": len(df),
                    "n_nonmissing_key": None,
                    "n_unique_key": None,
                    "n_duplicate_rows_by_key": None,
                    "status": "missing_key_column",
                }
            )
            return

        s = df[key].dropna().astype(str)
        n_rows = len(df)
        n_nonmissing = len(s)
        n_unique = s.nunique()
        n_dupe_rows = n_nonmissing - n_unique

        checks.append(
            {
                "table_name": table_name,
                "key_column": key,
                "exists": True,
                "n_rows": n_rows,
                "n_nonmissing_key": n_nonmissing,
                "n_unique_key": n_unique,
                "n_duplicate_rows_by_key": n_dupe_rows,
                "status": "ok" if n_dupe_rows == 0 else "duplicates_expected_or_problem",
            }
        )

    add_check("aiddata_records", "aiddata_record_id")
    add_check("aiddata_borrower_ownership", "aiddata_record_id")

    return pd.DataFrame(checks)


def build_relationship_audit(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []

    records = tables.get("aiddata_records")
    ownership = tables.get("aiddata_borrower_ownership")

    if records is None or ownership is None:
        return pd.DataFrame(
            [
                {
                    "relationship": "aiddata_records_to_borrower_ownership",
                    "status": "missing_required_table",
                }
            ]
        )

    if "aiddata_record_id" not in records.columns or "aiddata_record_id" not in ownership.columns:
        return pd.DataFrame(
            [
                {
                    "relationship": "aiddata_records_to_borrower_ownership",
                    "status": "missing_aiddata_record_id_column",
                }
            ]
        )

    parent_ids = set(records["aiddata_record_id"].dropna().astype(str))
    child_ids = ownership["aiddata_record_id"].dropna().astype(str)

    child_unique_ids = set(child_ids)
    orphan_ids = sorted(child_unique_ids - parent_ids)
    matched_child_rows = child_ids.isin(parent_ids).sum()
    orphan_child_rows = (~child_ids.isin(parent_ids)).sum()

    rows.append(
        {
            "relationship": "aiddata_records.aiddata_record_id -> aiddata_borrower_ownership.aiddata_record_id",
            "parent_table": "aiddata_records",
            "child_table": "aiddata_borrower_ownership",
            "parent_rows": len(records),
            "parent_unique_ids": len(parent_ids),
            "child_rows": len(ownership),
            "child_unique_ids": len(child_unique_ids),
            "matched_child_rows": int(matched_child_rows),
            "orphan_child_rows": int(orphan_child_rows),
            "orphan_unique_ids": len(orphan_ids),
            "example_orphan_ids": " | ".join(orphan_ids[:20]),
            "status": "ok" if orphan_child_rows == 0 else "has_orphans",
        }
    )

    # Parent records with borrower ownership.
    rows.append(
        {
            "relationship": "aiddata_records with borrower ownership",
            "parent_table": "aiddata_records",
            "child_table": "aiddata_borrower_ownership",
            "parent_rows": len(records),
            "parent_unique_ids": len(parent_ids),
            "child_rows": len(ownership),
            "child_unique_ids": len(child_unique_ids),
            "matched_child_rows": None,
            "orphan_child_rows": None,
            "orphan_unique_ids": None,
            "example_orphan_ids": "",
            "status": f"{len(parent_ids & child_unique_ids)} parent ids have borrower ownership rows",
        }
    )

    return pd.DataFrame(rows)


def build_column_coverage(tables: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []

    for table_name, df in tables.items():
        n = len(df)
        for col in df.columns:
            s = df[col]
            nonmissing = s.notna() & (s.astype(str).str.strip() != "")
            examples = (
                s[nonmissing]
                .astype(str)
                .drop_duplicates()
                .head(5)
                .tolist()
            )

            rows.append(
                {
                    "table_name": table_name,
                    "column_name": col,
                    "n_rows": n,
                    "n_nonmissing": int(nonmissing.sum()),
                    "nonmissing_share": round(float(nonmissing.mean()), 4) if n else None,
                    "n_unique": int(s[nonmissing].nunique()),
                    "examples": " | ".join(examples)[:700],
                }
            )

    return pd.DataFrame(rows)


def write_readme(
    outdir: Path,
    input_path: Path,
    table_inventory: pd.DataFrame,
) -> None:
    lines = []
    lines.append("# AidData CLG-LMIC v1.0 relational CSV extraction\n")
    lines.append(f"Created at: `{datetime.utcnow().isoformat()}Z`\n")
    lines.append(f"Input workbook: `{input_path}`\n")
    lines.append("\n## Tables\n")

    for _, row in table_inventory.iterrows():
        lines.append(f"- `{row['table_name']}.csv`: {row['n_rows']} rows, {row['n_cols']} columns")

    lines.append(
        """

## Notes

- Raw workbook was not modified.
- CSV column names were normalized to snake_case.
- Original-to-clean column mapping is in `audit/column_name_mapping.csv`.
- Main parent table is `tables/aiddata_records.csv`.
- Borrower ownership is a child table linked by `aiddata_record_id`.
- Documentation sheets (`Contents`, `Overview`, `Guide`, `License`) are not exported as analytical tables.
"""
    )

    (outdir / "README.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        required=True,
        help="Path to AidData CLG-LMIC Excel workbook",
    )
    parser.add_argument(
        "--outdir",
        default="data/interim/aiddata_clg_lmic_relational_v1_0",
        help="Output directory for relational CSV extraction",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing output CSVs/audits",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.input)
    outdir = Path(args.outdir)
    out_tables = outdir / "tables"
    out_audit = outdir / "audit"

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {xlsx_path}")

    if outdir.exists() and not args.overwrite:
        raise SystemExit(
            f"Output directory already exists: {outdir}\n"
            f"Use --overwrite if you want to replace outputs."
        )

    out_tables.mkdir(parents=True, exist_ok=True)
    out_audit.mkdir(parents=True, exist_ok=True)

    print(f"Input workbook: {xlsx_path}")
    print(f"Output dir: {outdir}")

    print("Inspecting workbook sheets...")
    sheet_inventory = workbook_sheet_inventory(xlsx_path)
    sheet_inventory.to_csv(out_audit / "sheet_inventory.csv", index=False)

    exported_tables: Dict[str, pd.DataFrame] = {}
    mappings: List[pd.DataFrame] = []

    print("Extracting known analytical/reference sheets...")
    for sheet_name, meta in KNOWN_TABLES.items():
        table_name = meta["table_name"]
        header = meta["header"]

        print(f"  {sheet_name} -> {table_name} header={header}")
        df = read_known_sheet(xlsx_path, sheet_name=sheet_name, header=header)
        clean_df, mapping = export_table(
            df=df,
            table_name=table_name,
            sheet_name=sheet_name,
            out_tables=out_tables,
        )

        exported_tables[table_name] = clean_df
        mappings.append(mapping)

    print("Writing audits...")
    column_mapping = pd.concat(mappings, ignore_index=True)
    table_inventory = build_table_inventory(exported_tables)
    key_audit = build_key_uniqueness_audit(exported_tables)
    relationship_audit = build_relationship_audit(exported_tables)
    column_coverage = build_column_coverage(exported_tables)

    column_mapping.to_csv(out_audit / "column_name_mapping.csv", index=False)
    table_inventory.to_csv(out_audit / "table_inventory.csv", index=False)
    key_audit.to_csv(out_audit / "key_uniqueness_audit.csv", index=False)
    relationship_audit.to_csv(out_audit / "relationship_audit.csv", index=False)
    column_coverage.to_csv(out_audit / "column_coverage.csv", index=False)

    metadata = {
        "created_at_utc": datetime.utcnow().isoformat() + "Z",
        "input_workbook": str(xlsx_path),
        "output_directory": str(outdir),
        "known_tables": KNOWN_TABLES,
        "doc_sheets_skipped": sorted(DOC_SHEETS),
    }
    (out_audit / "extraction_metadata.json").write_text(
        json.dumps(metadata, indent=2),
        encoding="utf-8",
    )

    write_readme(outdir, xlsx_path, table_inventory)

    print("\nDone.")
    print(table_inventory.to_string(index=False))
    print("\nRelationship audit:")
    print(relationship_audit.to_string(index=False))


if __name__ == "__main__":
    main()
